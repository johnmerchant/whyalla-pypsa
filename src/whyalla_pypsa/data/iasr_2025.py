"""IASR 2025 workbook loaders — thin stubs.

Intentionally minimal: whyalla-pypsa only needs a few well-known numbers and
we do NOT depend on `isp-workbook-parser`. Callers pass an explicit workbook
path and a technology / table key.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - openpyxl is a hard dep at runtime
    load_workbook = None  # type: ignore[assignment]


# TODO: confirm actual sheet + column positions in the 2025 IASR workbook.
# The sheet names below are reasonable guesses; the loader returns a header
# echo rather than guessing at numeric columns.
_GENCOST_SHEET = "Build costs"


def load_gencost(workbook_path: str | Path, technology: str) -> dict[str, Any]:
    """Return header row (+ matching row if found) for a given technology.

    Pragmatic stub: reads the first sheet named `_GENCOST_SHEET`, scans column A
    for a row whose value matches `technology`, returns a dict of
    {header: cell} for that row. Raises FileNotFoundError / KeyError on miss.

    TODO: extend to return a structured cost dict once the exact IASR 2025
    sheet layout is confirmed (sheet name, header row, unit columns).
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for iasr_2025 loaders")
    wb = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    if _GENCOST_SHEET not in wb.sheetnames:
        raise KeyError(
            f"Sheet {_GENCOST_SHEET!r} not found in {workbook_path}; "
            f"available: {wb.sheetnames}"
        )
    ws = wb[_GENCOST_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    for row in rows:
        if row and row[0] == technology:
            return {str(h): v for h, v in zip(header, row)}
    raise KeyError(f"Technology {technology!r} not found in {_GENCOST_SHEET}")


# Placeholder REZ build limits (MW). Flagged loudly — these are NOT taken from
# the workbook yet. S4/S5/S6 are the Whyalla-adjacent SA REZs.
_FALLBACK_REZ_BUILD_LIMITS: dict[str, float] = {
    "S4": 2000.0,  # PLACEHOLDER
    "S5": 3000.0,  # PLACEHOLDER
    "S6": 2500.0,  # PLACEHOLDER
}


def load_rez_build_limits(workbook_path: str | Path | None = None) -> dict[str, float]:
    """Return REZ build limits (MW). Falls back to hand-coded placeholders.

    TODO: replace with direct read from IASR 2025 once the REZ build-limit
    sheet + column layout is confirmed. The numbers below are placeholders
    intended to be obviously wrong if mis-used.
    """
    # If caller didn't supply a workbook, just return placeholders.
    if workbook_path is None:
        return dict(_FALLBACK_REZ_BUILD_LIMITS)
    # Even with a path we currently return placeholders rather than guess at
    # the sheet layout.
    return dict(_FALLBACK_REZ_BUILD_LIMITS)
