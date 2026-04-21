"""Draft 2026 ISP Inputs and Assumptions workbook (IASR) reader.

Thin sibling to `iasr_2025.py`: exposes only the handful of well-known inputs
whyalla-pypsa needs. Minimal openpyxl scanning — no full workbook parser.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# Sheet names discovered in the 2026 draft workbook. Keep as module constants
# so they are easy to update when AEMO republishes.
_BUILD_COSTS_SHEET = "Build costs"
_FUEL_PRICE_SHEET = "Gas, Liquid fuel, H2 price"


def list_sheets(workbook_path: str | Path) -> list[str]:
    """Return sheet names in the IASR workbook (for discovery/debugging)."""
    wb = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    return list(wb.sheetnames)


def load_gencost(workbook_path: str | Path, technology: str) -> dict[str, Any]:
    """Return {header: cell} for the first row matching `technology` in the
    Build costs sheet. Raises KeyError on miss."""
    wb = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    if _BUILD_COSTS_SHEET not in wb.sheetnames:
        raise KeyError(
            f"Sheet {_BUILD_COSTS_SHEET!r} not found; available: {wb.sheetnames}"
        )
    ws = wb[_BUILD_COSTS_SHEET]

    # The sheet has a leading blank column and ~9 lead-in rows of commentary.
    # The real header row starts with 'Technology' in some form (the workbook
    # uses 'Technology1' with a footnote suffix).
    header: tuple[Any, ...] | None = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and any(
                isinstance(c, str) and c.startswith("Technology") for c in row
            ):
                header = row
            continue
        if not row:
            continue
        # Technology is whichever column the header row marks; find its index.
        tech_idx = next(
            (i for i, h in enumerate(header) if isinstance(h, str) and h.startswith("Technology")),
            None,
        )
        if tech_idx is None:
            break
        if row[tech_idx] == technology:
            return {str(h): v for h, v in zip(header, row) if h is not None}
    raise KeyError(f"Technology {technology!r} not found in {_BUILD_COSTS_SHEET}")


def load_fuel_price(workbook_path: str | Path, fuel: str, fy: int) -> float:
    """Return a fuel price ($/GJ) for a given fuel identifier and FY.

    The 2026 fuel price sheet is scenario-structured (per-generator × gas-price
    scenario) rather than a clean per-fuel table, so the exact key semantics are
    not fully pinned down. The function raises a descriptive error rather than
    guess — see `load_gencost` for a working example.
    """
    raise NotImplementedError(
        f"load_fuel_price not yet implemented for 2026 IASR layout "
        f"(requested fuel={fuel!r}, fy={fy}). "
        f"See sheet {_FUEL_PRICE_SHEET!r} for the generator×scenario table."
    )
