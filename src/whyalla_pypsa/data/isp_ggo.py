"""Reader for AEMO Draft 2026 ISP Generation & Storage Outlook (GGO) Cores.

Exposes the ODP capacity trajectory per NEM subregion × technology × year so
the Whyalla facility model can be layered on top of AEMO's solved background.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


# The CDP column uses the literal string 'CDP4 (ODP)' with parenthetical — a
# plain 'CDP4' filter returns zero rows. Keep this string exact.
ODP_CDP = "CDP4 (ODP)"


def _fy_from_label(label: Any) -> int | None:
    """Parse '2029-30' -> 2030. Returns None if not a FY label."""
    if not isinstance(label, str):
        return None
    parts = label.split("-")
    if len(parts) != 2:
        return None
    try:
        return 2000 + int(parts[1])
    except ValueError:
        return None


def load_ggo_capacity(
    workbook_path: str | Path,
    *,
    cdp: str = ODP_CDP,
    subregion: str | None = None,
    technology: str | None = None,
    sheet: str = "Capacity",
) -> pd.DataFrame:
    """Return installed capacity (MW) in long format for the GGO workbook.

    Columns: ['cdp', 'subregion', 'technology', 'fy', 'capacity_mw'].
    `sheet` selects among 'Capacity' (generation), 'Storage Capacity',
    'Electrolyzer Capacity'; the last has a 5-ID-column layout.
    """
    wb = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet!r} not found; available: {wb.sheetnames}")
    ws = wb[sheet]

    header: tuple[Any, ...] | None = None
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "CDP":
            header = row
            continue
        if header is None or not row or not row[0]:
            continue
        if row[0] != cdp:
            continue

        # Identify where year columns begin by scanning the header.
        year_cols = [(i, _fy_from_label(h)) for i, h in enumerate(header)]
        id_stop = next(i for i, fy in year_cols if fy is not None)
        # Subregion is column 2 in every sheet (CDP/Region/Subregion/...).
        sub = row[2]
        tech = row[id_stop - 1]  # last ID column before the years
        if subregion is not None and sub != subregion:
            continue
        if technology is not None and tech != technology:
            continue

        for i, fy in year_cols:
            if fy is None:
                continue
            val = row[i]
            if val is None:
                continue
            records.append(
                {
                    "cdp": row[0],
                    "subregion": sub,
                    "technology": tech,
                    "fy": fy,
                    "capacity_mw": float(val),
                }
            )

    df = pd.DataFrame(records, columns=["cdp", "subregion", "technology", "fy", "capacity_mw"])
    if not df.empty:
        df = df.astype({"fy": "int64"})
    return df


def list_sheets(workbook_path: str | Path) -> list[str]:
    """Return sheet names in a GGO Cores workbook (for discovery/debugging)."""
    wb = load_workbook(Path(workbook_path), read_only=True, data_only=True)
    return list(wb.sheetnames)
